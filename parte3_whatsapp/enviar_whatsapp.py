import time, os, sys, csv, pyautogui, urllib.parse, openpyxl, copy
from datetime import datetime
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from dotenv import load_dotenv

load_dotenv("config/.env")
ESPERA_WHATSAPP = int(os.getenv("ESPERA_WHATSAPP", 15))
URL_SISTEMA     = os.getenv("URL_SISTEMA", "http://127.0.0.1:5000")
NOME_EMPRESA    = os.getenv("NOME_EMPRESA", "TechSolutions Ltda")
ARQUIVO_ERROS   = "dados/erros.csv"
ARQUIVO_RELATORIO = "dados/faturas_relatorio.xlsx"
PASTA_BOLETOS   = "boletos"

sys.path.insert(0, "1parte_web")
from database import listar_faturas_pendentes, atualizar_status_fatura

def validar_telefone(tel):
    limpo = "".join(c for c in str(tel) if c.isdigit())
    return len(limpo) >= 12, limpo

def montar_mensagem(fatura):
    nome_arq    = f"boleto_cliente{fatura['id']:03d}_{fatura['nome'].replace(' ','_')}.pdf"
    link_boleto = f"{URL_SISTEMA}/boletos/{nome_arq}"
    return (
        f"Olá, *{fatura['nome']}*! 👋\n\n"
        f"Notificação automática da *{NOME_EMPRESA}*.\n\n"
        f"📄 *Fatura #{fatura['id']:05d}*\n"
        f"💰 Valor: *R$ {fatura['valor']:.2f}*\n"
        f"📅 Vencimento: *{fatura['data_vencimento']}*\n\n"
        f"Pague pelo QR Code PIX:\n🔗 {link_boleto}\n\n"
        f"Dúvidas? Responda esta mensagem.\n"
        f"_{NOME_EMPRESA} — Financeiro_ 🏢"
    )

def registrar_erro(fatura, erro):
    os.makedirs("dados", exist_ok=True)
    existe = os.path.exists(ARQUIVO_ERROS)
    with open(ARQUIVO_ERROS, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if not existe:
            w.writerow(["fatura_id", "cliente", "erro", "data_hora"])
        w.writerow([fatura["id"], fatura["nome"], erro,
                    datetime.now().strftime("%d/%m/%Y %H:%M:%S")])

def enviar_whatsapp(driver, fatura):
    valido, telefone = validar_telefone(fatura["telefone"])
    if not valido:
        raise ValueError(f"Telefone inválido: {fatura['telefone']}")

    url = f"https://web.whatsapp.com/send?phone={telefone}&text={urllib.parse.quote(montar_mensagem(fatura))}"
    print(f"   🌐 Abrindo WhatsApp para {fatura['nome']} ({telefone})...")
    driver.get(url)

    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.webdriver.common.keys import Keys

    wait = WebDriverWait(driver, 40)
    try:
        # Aguarda a tela principal do chat carregar
        wait.until(EC.presence_of_element_located((By.ID, 'main')))
        time.sleep(3) # Aguarda a mensagem ser preenchida na caixa de texto
        
        # Tenta clicar no botão de enviar primeiramente
        try:
            btn_enviar = driver.find_element(By.XPATH, '//button[@aria-label="Enviar"] | //span[@data-icon="send"]/..')
            btn_enviar.click()
        except:
            # Fallback seguro: enviar tecla ENTER pela própria janela do navegador
            ActionChains(driver).send_keys(Keys.ENTER).perform()
            
    except Exception as e:
        print("   ⚠️ Aviso: Não localizamos o chat, usando fallback de Enter via PyAutoGUI...")
        time.sleep(2)
        pyautogui.hotkey("enter")

    time.sleep(2)
    print(f"   ✅ Enviado para {fatura['nome']}!")

def ler_faturas_excel():
    if not os.path.exists(ARQUIVO_RELATORIO):
        print(f"⚠️ Arquivo {ARQUIVO_RELATORIO} não encontrado. Gere o relatório primeiro!")
        return [], None, None, None, None
    
    wb = openpyxl.load_workbook(ARQUIVO_RELATORIO)
    if "Faturas" not in wb.sheetnames:
        print(f"⚠️ Aba 'Faturas' não encontrada em {ARQUIVO_RELATORIO}.")
        return [], wb, None, None, None

    ws = wb["Faturas"]
    
    colunas = {}
    for col in range(1, ws.max_column + 2):
        val = ws.cell(row=3, column=col).value
        if val:
            colunas[val] = col
            
    if "Status" not in colunas:
        print("⚠️ Coluna 'Status' não encontrada na planilha.")
        return [], wb, ws, None, None
        
    faturas = []
    for row in range(4, ws.max_row + 1):
        status = ws.cell(row=row, column=colunas["Status"]).value
        if status and str(status).strip().lower() == "pendente":
            try:
                data_vencimento = ws.cell(row=row, column=colunas.get("Vencimento", 6)).value
                if isinstance(data_vencimento, datetime):
                    data_vencimento = data_vencimento.strftime("%Y-%m-%d")
                    
                fatura = {
                    "id": ws.cell(row=row, column=colunas.get("ID Fatura", 1)).value,
                    "nome": ws.cell(row=row, column=colunas.get("Cliente", 2)).value,
                    "email": ws.cell(row=row, column=colunas.get("E-mail", 3)).value,
                    "telefone": ws.cell(row=row, column=colunas.get("Telefone", 4)).value,
                    "valor": ws.cell(row=row, column=colunas.get("Valor (R$)", 5)).value,
                    "data_vencimento": data_vencimento,
                    "row_idx": row 
                }
                faturas.append(fatura)
            except Exception as e:
                print(f"⚠️ Erro ao ler fatura na linha {row}: {e}")
            
    if "Data/Hora Envio" not in colunas:
        nova_col = ws.max_column + 1
        ws.cell(row=3, column=nova_col, value="Data/Hora Envio")
        header_base = ws.cell(row=3, column=colunas["Status"])
        if header_base.has_style:
            ws.cell(row=3, column=nova_col).font = copy.copy(header_base.font)
            ws.cell(row=3, column=nova_col).fill = copy.copy(header_base.fill)
            ws.cell(row=3, column=nova_col).border = copy.copy(header_base.border)
            ws.cell(row=3, column=nova_col).alignment = copy.copy(header_base.alignment)
        colunas["Data/Hora Envio"] = nova_col
        ws.column_dimensions[openpyxl.utils.get_column_letter(nova_col)].width = 20

    return faturas, wb, ws, colunas["Status"], colunas["Data/Hora Envio"]

def atualizar_excel_status(wb, ws, row_idx, col_status, col_data, novo_status, caminho):
    verde    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    amarelo  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    cell_status = ws.cell(row=row_idx, column=col_status)
    cell_status.value = novo_status.capitalize()
    
    cell_data = ws.cell(row=row_idx, column=col_data)
    cell_data.value = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    fmt = verde if novo_status.lower() == "enviado" else (vermelho if novo_status.lower() == "falhou" else amarelo)
    
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row_idx, column=col).fill = fmt
        
    wb.save(caminho)

def main():
    print("=" * 50)
    print("  📱 RPA - Envio via WhatsApp Web (Integrado com Excel)")
    print("=" * 50)

    faturas, wb, ws, col_status, col_data = ler_faturas_excel()
    if not faturas:
        print("⚠️  Nenhuma fatura pendente na planilha (ou arquivo vazio).")
        return

    print(f"\n📋 {len(faturas)} fatura(s) pendente(s) lida(s) da planilha Excel.\n")

    opcoes = webdriver.ChromeOptions()
    opcoes.add_argument("--start-maximized")
    # Descomente para manter sessão entre execuções:
    # opcoes.add_argument("--user-data-dir=./whatsapp_session")

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=opcoes
    )

    try:
        print(f"📱 Abrindo WhatsApp Web...")
        print(f"   ⏳ Você tem {ESPERA_WHATSAPP}s para escanear o QR Code!")
        driver.get("https://web.whatsapp.com")
        time.sleep(ESPERA_WHATSAPP)

        enviados = falhas = 0
        for i, f in enumerate(faturas, 1):
            print(f"\n[{i}/{len(faturas)}] {f['nome']} — R$ {f['valor']:.2f}")
            try:
                enviar_whatsapp(driver, f)
                # Atualizar DADOS NO BANCO para sincronia
                atualizar_status_fatura(f["id"], "enviado")
                # Atualizar EXCEL (Parte 5)
                atualizar_excel_status(wb, ws, f["row_idx"], col_status, col_data, "Enviado", ARQUIVO_RELATORIO)
                enviados += 1
            except Exception as e:
                print(f"   ❌ ERRO: {e}")
                registrar_erro(f, str(e))
                # Atualizar DADOS NO BANCO para sincronia
                atualizar_status_fatura(f["id"], "falhou")
                # Atualizar EXCEL (Parte 5)
                atualizar_excel_status(wb, ws, f["row_idx"], col_status, col_data, "Falhou", ARQUIVO_RELATORIO)
                falhas += 1
            time.sleep(3)
    finally:
        driver.quit()

    print(f"\n{'='*50}")
    print(f"✅ Enviados: {enviados} | ❌ Falhas: {falhas}")
    if falhas: print(f"📄 Erros em: {ARQUIVO_ERROS}")
    print(f"📊 Planilha atualizada salva em: {ARQUIVO_RELATORIO}")

if __name__ == "__main__":
    main()