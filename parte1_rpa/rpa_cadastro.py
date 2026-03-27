import time, openpyxl, sys, os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from dotenv import load_dotenv

load_dotenv("config/.env")
URL_SISTEMA      = os.getenv("URL_SISTEMA", "http://127.0.0.1:5000")
ARQUIVO_CLIENTES = "dados/clientes_importar.xlsx"

def ler_clientes_excel(caminho):
    print(f"\n📊 Lendo: {caminho}")
    wb = openpyxl.load_workbook(caminho)
    ws = wb.active
    clientes = []
    for linha in ws.iter_rows(min_row=2, values_only=True):
        if linha[0]:
            clientes.append({
                "nome":     str(linha[0]).strip(),
                "email":    str(linha[1]).strip(),
                "telefone": str(linha[2]).strip(),
                "endereco": str(linha[3]).strip() if linha[3] else "",
            })
    print(f"✅ {len(clientes)} clientes encontrados")
    return clientes, wb, ws

def iniciar_navegador():
    opcoes = webdriver.ChromeOptions()
    opcoes.add_argument("--start-maximized")
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=opcoes
    )

def cadastrar_cliente_selenium(driver, cliente):
    wait = WebDriverWait(driver, 10)
    driver.get(f"{URL_SISTEMA}/clientes/novo")
    time.sleep(1)

    wait.until(EC.presence_of_element_located((By.NAME, "nome"))).clear()
    driver.find_element(By.NAME, "nome").send_keys(cliente["nome"])
    driver.find_element(By.NAME, "email").send_keys(cliente["email"])
    driver.find_element(By.NAME, "telefone").send_keys(cliente["telefone"])
    driver.find_element(By.NAME, "endereco").send_keys(cliente["endereco"])
    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
    time.sleep(1.5)

    try:
        alerta = driver.find_element(By.CLASS_NAME, "alerta")
        return "sucesso" in alerta.get_attribute("class")
    except:
        return True

def atualizar_excel(caminho, wb, ws, resultados):
    from openpyxl.styles import PatternFill, Font
    verde    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    col      = ws.max_column + 1
    ws.cell(row=1, column=col, value="Status RPA").font = Font(bold=True)
    for i, ok in enumerate(resultados, start=2):
        cell       = ws.cell(row=i, column=col)
        cell.value = "✅ Cadastrado" if ok else "❌ Falhou"
        cell.fill  = verde if ok else vermelho
    wb.save(caminho)
    print(f"\n📊 Planilha atualizada: {caminho}")

def main():
    print("=" * 50)
    print("  🤖 RPA - Cadastro Automático de Clientes")
    print("=" * 50)

    if not os.path.exists(ARQUIVO_CLIENTES):
        print(f"❌ Arquivo não encontrado: {ARQUIVO_CLIENTES}")
        print("   Crie o Excel com colunas: Nome | Email | Telefone | Endereco")
        return

    clientes, wb, ws = ler_clientes_excel(ARQUIVO_CLIENTES)
    if not clientes:
        print("❌ Nenhum cliente na planilha!")
        return

    print("\n🌐 Iniciando Chrome...")
    driver     = iniciar_navegador()
    resultados = []

    try:
        for i, c in enumerate(clientes, 1):
            print(f"\n[{i}/{len(clientes)}] Cadastrando: {c['nome']}...")
            ok = cadastrar_cliente_selenium(driver, c)
            print(f"   {'✅ Sucesso!' if ok else '❌ Falhou!'}")
            resultados.append(ok)
            time.sleep(1)
    finally:
        driver.quit()

    atualizar_excel(ARQUIVO_CLIENTES, wb, ws, resultados)
    print(f"\n📊 RESUMO: {sum(resultados)}/{len(resultados)} cadastrados!")

if __name__ == "__main__":
    main()