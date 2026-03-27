import os, sys
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv("config/.env")
ARQUIVO_RELATORIO = "dados/faturas_relatorio.xlsx"
NOME_EMPRESA      = os.getenv("NOME_EMPRESA", "TechSolutions Ltda")

sys.path.insert(0, "1parte_web")
from database import conectar

def importar_dados_banco():
    conn = conectar()
    df = pd.read_sql_query("""
        SELECT f.id AS "ID Fatura", c.nome AS "Cliente", c.email AS "E-mail",
               c.telefone AS "Telefone", f.valor AS "Valor (R$)",
               f.data_vencimento AS "Vencimento", f.data_emissao AS "Emissão",
               f.status AS "Status"
        FROM faturas f
        JOIN clientes c ON f.cliente_id = c.id
        ORDER BY f.data_vencimento
    """, conn)
    conn.close()
    print(f"✅ {len(df)} registros importados do banco")
    return df

def exportar_relatorio_excel(df, caminho):
    os.makedirs("dados", exist_ok=True)

    with pd.ExcelWriter(caminho, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Faturas", index=False, startrow=2)
        wb  = writer.book
        ws  = writer.sheets["Faturas"]

        f_titulo  = wb.add_format({"bold":True,"font_size":14,"font_color":"#FFFFFF",
                                    "bg_color":"#1a237e","align":"center","valign":"vcenter"})
        f_verde   = wb.add_format({"bg_color":"#C6EFCE","font_color":"#276221"})
        f_vermelho= wb.add_format({"bg_color":"#FFC7CE","font_color":"#9C0006"})
        f_amarelo = wb.add_format({"bg_color":"#FFEB9C","font_color":"#9C6500"})

        ws.merge_range("A1:H1",
            f"{NOME_EMPRESA} — Relatório de Faturas — {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            f_titulo)
        ws.set_row(0, 30)

        for col, larg in enumerate([10,25,30,18,14,12,12,12]):
            ws.set_column(col, col, larg)

        for i in range(len(df)):
            status = str(df.iloc[i]["Status"]).lower()
            fmt    = f_verde if status=="enviado" else (f_vermelho if status=="falhou" else f_amarelo)
            ws.set_row(i + 3, None, fmt)

    # Aba resumo
    wb2 = openpyxl.load_workbook(caminho)
    _aba_resumo(wb2, df)
    wb2.save(caminho)
    print(f"✅ Relatório salvo: {caminho}")

def _aba_resumo(wb, df):
    if "Resumo" in wb.sheetnames:
        del wb["Resumo"]
    ws  = wb.create_sheet("Resumo")
    azul = PatternFill("solid", fgColor="1a237e")
    bnco = Font(color="FFFFFF", bold=True)
    brd  = Border(left=Side(style="thin"),right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    ws.merge_cells("A1:C1")
    ws["A1"] = f"Resumo — {datetime.now().strftime('%d/%m/%Y')}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = azul
    ws["A1"].alignment = Alignment(horizontal="center")

    for col, titulo in enumerate(["Status","Qtd. Faturas","Valor Total (R$)"], 1):
        c = ws.cell(1+1, col, titulo)
        c.font = bnco
        c.fill = PatternFill("solid", fgColor="283593")
        c.border = brd

    resumo = df.groupby("Status").agg(
        Qtd=("ID Fatura","count"), Total=("Valor (R$)","sum")
    ).reset_index()
    cores = {"enviado":"C6EFCE","pendente":"FFEB9C","falhou":"FFC7CE"}

    for i, row in resumo.iterrows():
        l = i + 3
        ws.cell(l,1,row["Status"].upper())
        ws.cell(l,2,row["Qtd"])
        ws.cell(l,3,round(row["Total"],2))
        for c in range(1,4):
            ws.cell(l,c).fill = PatternFill("solid", fgColor=cores.get(row["Status"].lower(),"FFFFFF"))
            ws.cell(l,c).border = brd

    ul = len(resumo)+3
    for c, v in zip(range(1,4), ["TOTAL", len(df), round(df["Valor (R$)"].sum(),2)]):
        ws.cell(ul,c,v).font = Font(bold=True)
        ws.cell(ul,c).border = brd

    for c, w in zip("ABC",[15,18,20]):
        ws.column_dimensions[c].width = w

def main():
    print("=" * 50)
    print("  📊 RPA - Relatório Excel")
    print("=" * 50)

    print("\n1️⃣  Importando dados do banco...")
    df = importar_dados_banco()
    if df.empty:
        print("⚠️  Banco vazio. Cadastre clientes/faturas primeiro!")
        return

    print("\n2️⃣  Gerando planilha Excel...")
    exportar_relatorio_excel(df, ARQUIVO_RELATORIO)

    print("\n3️⃣  Resumo:")
    print("-" * 40)
    for status, g in df.groupby("Status"):
        emoji = "✅" if status=="enviado" else ("❌" if status=="falhou" else "⏳")
        print(f"   {emoji} {status.upper():12} | {len(g):3} fatura(s) | R$ {g['Valor (R$)'].sum():.2f}")
    print("-" * 40)
    print(f"   📊 TOTAL         | {len(df):3} fatura(s) | R$ {df['Valor (R$)'].sum():.2f}")
    print(f"\n✅ Arquivo: {ARQUIVO_RELATORIO}")

if __name__ == "__main__":
    main()